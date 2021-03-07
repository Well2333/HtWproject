from PySide2.QtWidgets import *
from PySide2.QtUiTools import *
import win32com.client as win32
import openpyxl as xl
import os , csv
#获取工作路径,判断input与output是否存在
path = os.getcwd()
if not os.path.exists(f"{path}\\input"):
    os.makedirs(f"{path}\\input")
if not os.path.exists(f"{path}\\output"):
    os.makedirs(f"{path}\\output")

class main:
    Files = []
    def __init__(self):
        self.ui = QUiLoader().load('ui/main.ui')
        self.ui.ScheduleCheck.clicked.connect(self.check_file)
        self.ui.ScheduleTrans.clicked.connect(self.file_trans)
    #执行课表转换
    def file_trans(self):
        self.ui.ErrorInfo.appendPlainText("Info:开始转换课表")
        self.ui.progressBar.setRange(0,len(self.Files))
        for frow in range(0,len(self.Files)):
            #转换课表格式
            transapp = win32.gencache.EnsureDispatch('Excel.Application')
            filetrans = transapp.Workbooks.Open(f"{path}\\input\\{self.Files[frow]}")
            filetrans.SaveAs(f"{path}\\tempfile.xlsx", FileFormat=51)
            filetrans.Close()      
            transapp.Application.Quit()                         
            #读取课表内容
            filename_og,_ = self.Files[frow].split(".")
            self.FileTrans(f"{filename_og}.csv")
            os.remove("tempfile.xlsx")
            self.ui.progressBar.setValue(frow+1)
        self.ui.ErrorInfo.appendPlainText("Info:课表已转换完成，请在output中查收")
    #读取input中的文件
    def check_file(self):
        #重置界面
        self.ui.FileList.clear()
        self.Files = []
        #按照类别读取输入文件，并尝试自动补全信息
        FilePath_str = os.listdir(path + "\\" + "input")
        for File in FilePath_str:
            if ".xls" in File:
                self.Files.append(File)
        if len(self.Files) == 0:
            self.ui.ErrorInfo.appendPlainText("Warning:请在input文件夹中放入课表源文件")
        else:
            self.ui.FileList.addItems(self.Files)
            self.ui.ErrorInfo.appendPlainText("Info:课表已检索完成")
            self.ui.ErrorInfo.appendPlainText("Guide:请核对课表名称是否正确，然后点击转换课表")
            self.ui.ScheduleTrans.setEnabled(True)
    #转换课表的具体操作
    def FileTrans(self,filename):
        #读取tempfile文件
        wb_in = xl.load_workbook("tempfile.xlsx",data_only=True)
        ws_in = wb_in["Sheet1"]
        #检查是否有重名输出
        files = os.listdir(f"{path}\\output")
        for f in files:
            if filename in f:
                self.ui.ErrorInfo.appendPlainText(f'Info:{filename}已经存在，进行覆盖')
                os.remove(f"{path}\\output\\{filename}")
        #生成输出文件
        f = open(f"{path}\\output\\{filename}",'w',encoding="utf-8",newline='' "")
        csv_writer = csv.writer(f)
        csv_writer.writerow(["课程名称","星期","开始节数","结束节数","老师","地点","周数"])
        #遍历课程
        for row_num in range(1,ws_in.max_row):
            for column_num in range(1,ws_in.max_column):
                if ws_in.cell(row_num,column_num).value == ws_in.cell(500,500).value:
                    pass
                else:
                    #获取单元格名称
                    char = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
                    cellname = f"{char[column_num-1]}{row_num}"
                    #生成并填写至输出
                    try:
                        Infostr = str(ws_in[cellname].value).split('''_x000D_\n''')
                        classname = Infostr[1]
                        teacher = Infostr[2]
                        weeknum,timeinfo = Infostr[3].split("-")
                        start = int(timeinfo[:2])
                        end = int(timeinfo[-2:])
                        feild = Infostr[4]
                        weekday = (column_num-1)%7
                        if weekday == 0:
                            weekday = 7
                        csv_writer.writerow([classname,weekday,start,end,teacher,feild,weeknum])
                    except:
                        pass
        self.ui.ErrorInfo.appendPlainText(f"Info:{filename}已转换完成")
#开启界面
app = QApplication([])
Main = main()
Main.ui.show()
app.exec_()
#清理残余tempfile.xlsx
filec = os.listdir(path)
for fc in filec:
    if "tempfile.xlsx" in fc:
        os.remove(path+'\\'+"tempfile.xlsx")
        Main.ui.ErrorInfo.appendPlainText('Info:已清除旧的“tempfile.xlsx”')