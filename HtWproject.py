import openpyxl as xl
import win32com.client as win32
import csv
import os

print("开发者@Well404，如有bug或建议请加QQ1070330078反馈并说明来意。")
print("正在读取本文件夹内是否含有课表")

try:
    #获取原文件并转换
    path = os.getcwd()
    file = os.listdir(path)
    for f in file:
        if ".xls" in f:
            print("已检测到课表:D！正在转换为input.xlsx！")
            transapp = win32.gencache.EnsureDispatch('Excel.Application')
            filetrans = transapp.Workbooks.Open(path+'\\'+f)
            filetrans.SaveAs(path+'\\'+"input.xlsx", FileFormat=51)
            filetrans.Close()                               
            transapp.Application.Quit()
            print("成功转换课表:D！")
            break
        else:
            print("未检测到课表T_T，请将课表复制到于本文件相同的文件夹中并重启程序")

    class_number = input("请输入班级名(需要和表内的班级名一致):")
    print("正在转换，请稍等片刻OvO")

    #获取转化格式后的文件
    wb_in = xl.load_workbook("input.xlsx",data_only=True)
    ws_in = wb_in.active

    #执行拆分填充
    m_list = ws_in.merged_cells
    while len(ws_in.merged_cells.ranges) > 0:
        for m_area in m_list:
            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
            ws_in.unmerge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
            for rowm in range(r1,r2+1):
                for colm in range(c1,c2+1):
                    ws_in.cell(rowm,colm,value=ws_in.cell(r1,c1).value)

    #创建输出文件
    f = open('我是导入课表文件.csv','w',encoding="utf-8",newline='' "")
    csv_writer = csv.writer(f)
    csv_writer.writerow(["课程名称","星期","开始节数","结束节数","老师","地点","周数"])

    #获取week
    def week_num(row_num):
        if "本科临床医学专业" in ws_in.cell(1,1).value:
            number = 0
            for row_num_week in range(1,row_num):
                if str(ws_in.cell(row_num_week,3).value) == "班":
                    number = number + 1
                    return number//3
        else:
            number = 1
            while 1>0:
                if "第{0}周".format(number) in str(ws_in.cell(row_num,1).value):
                    return number
                    break
                else:
                    number = number + 1
    

    #获取class_num
    def class_num(column_num):
        if str(column_num) in ["4","12","20","28","36","44","52"]:
            return 1
        elif str(column_num) in ["5","13","21","29","37","45","53"]:
            return 2
        elif str(column_num) in ["6","14","22","30","38","46","54"]:
            return 3
        elif str(column_num) in ["7","15","23","31","39","47","55"]:
            return 4
        elif str(column_num) in ["8","16","24","32","40","48","56"]:
            return 5
        elif str(column_num) in ["9","17","25","33","41","49","57"]:
            return 6
        elif str(column_num) in ["10","18","26","34","42","50","58"]:
            return 7
        elif str(column_num) in ["11","19","27","35","43","51","59"]:
            return 8

    #获取class_name,start,end
    def class_name_str(row_num,column_num):
        return str(ws_in.cell(row_num,column_num).value)

    def start_num(column_num):
        return class_num(column_num)

    def end_num(column_num):
        if start_num(column_num) == 7:
            return 8
        elif class_name_str(row_num,column_num) == class_name_str(row_num,column_num+2):
            return class_num(column_num) + 3
        else:
            return class_num(column_num) + 1

    #获取week_day
    def week_day_num(column_num):
        if column_num < 12:
            return 1
        elif column_num < 20:
            return 2
        elif column_num < 28:
            return 3
        elif column_num < 36:
            return 4
        elif column_num < 44:
            return 5
        elif column_num < 52:
            return 6
        elif column_num < 60:
            return 7

    for row_num in range(1,ws_in.max_row):
        if class_number == str(ws_in.cell(row_num,2).value):
            for column_num in range(4,60,2):
                if class_name_str(row_num,column_num) == class_name_str(row_num,column_num-2) and class_num(column_num) > 1:
                    pass
                else:
                    if str(class_name_str(row_num,column_num)) == "None":
                        pass
                    else:
                        class_name = class_name_str(row_num,column_num)
                        week_day = week_day_num(column_num)
                        start = start_num(column_num)
                        end = end_num(column_num)
                        week = week_num(row_num)
                        csv_writer.writerow([class_name,week_day,start,end,"","",week])
    f.close()
    print("转换完成:D，快去导入课表吧！")
except:
    error_count = 10
    if error_count >0:
        print("程序运行异常X_X，请联系开发者寻求帮助！！！")
        error_count = error_count-1
os.remove(path+'\\'+"input.xlsx")
#os.remove(path+'\\'+f)
input("按任意键以退出程序:D")
